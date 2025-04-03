#!/usr/bin/env python3
"""
Email Module

This module provides functionality to:
1. Select a work week
2. Check for loadsheets and timesheet in the appropriate folders
3. Convert Excel files to PDF using LibreOffice
4. Organize PDFs in an email folder
5. Send email with PDFs and weekly summary
"""

import os
import logging
import shutil
import subprocess
import configparser
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from datetime import datetime, timedelta
import colorama
from colorama import Fore, Style
import psycopg2
from psycopg2.extras import RealDictCursor
import openpyxl

# Initialize colorama for cross-platform terminal colors
colorama.init(autoreset=True)

# -------- Logging Setup --------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LOG_DIR = os.path.join(SCRIPT_DIR, "logs")
os.makedirs(LOG_DIR, exist_ok=True)

# Setup logging configuration
LOG_FILE = os.path.join(LOG_DIR, f"email_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE)
    ]
)

def load_email_config():
    """Load email configuration from config file."""
    config = configparser.ConfigParser()
    config_path = os.path.join(SCRIPT_DIR, "config", "email_config.ini")
    
    if not os.path.exists(config_path):
        logging.error(f"Email configuration file not found: {config_path}")
        return None
    
    config.read(config_path)
    return config

def test_smtp_connection():
    """Test SMTP connection and credentials."""
    try:
        config = load_email_config()
        if not config:
            print(f"{Fore.RED}Email configuration not found. Please set up email_config.ini{Style.RESET_ALL}")
            return False
        
        print(f"\n{Fore.CYAN}Testing SMTP connection...{Style.RESET_ALL}")
        print(f"{Fore.WHITE}Server: {config['Email']['smtp_server']}")
        print(f"{Fore.WHITE}Port: {config['Email']['smtp_port']}")
        print(f"{Fore.WHITE}Email: {config['Email']['sender_email']}")
        
        with smtplib.SMTP(config['Email']['smtp_server'], 
                         int(config['Email']['smtp_port'])) as server:
            server.starttls()
            server.login(config['Email']['sender_email'], 
                        config['Email']['sender_password'])
            print(f"{Fore.GREEN}✓ SMTP connection successful!{Style.RESET_ALL}")
            return True
            
    except Exception as e:
        print(f"{Fore.RED}✗ SMTP connection failed: {str(e)}{Style.RESET_ALL}")
        return False

def get_week_summary(week_end_date, timesheet_file):
    """Get summary of cars for the week from timesheet Excel file."""
    try:
        # Load the timesheet workbook
        workbook = openpyxl.load_workbook(timesheet_file)
        sheet = workbook.active
        
        # Initialize counters
        total_loads = 0
        total_vehicles = 0
        
        # Start from row 8 (where load data begins)
        for row in range(8, sheet.max_row + 1):
            # Check if cell D has a value (number of cars)
            cars_cell = sheet.cell(row=row, column=4)  # Column D
            if cars_cell.value is not None:
                try:
                    cars = int(cars_cell.value)
                    total_loads += 1
                    total_vehicles += cars
                except (ValueError, TypeError):
                    continue
        
        workbook.close()
        
        # Create summary dictionary
        summary = {
            'total_loads': total_loads,
            'total_vehicles': total_vehicles
        }
        
        print(f"{Fore.GREEN}✓ Successfully read timesheet!{Style.RESET_ALL}")
        if total_loads > 0:
            print(f"{Fore.WHITE}Total Loads: {total_loads}")
            print(f"{Fore.WHITE}Total Vehicles: {total_vehicles}")
        else:
            print(f"{Fore.YELLOW}No loads found in timesheet - sending timesheet only{Style.RESET_ALL}")
        
        return summary
        
    except Exception as e:
        logging.error(f"Error reading timesheet: {e}")
        print(f"{Fore.RED}✗ Failed to read timesheet: {str(e)}{Style.RESET_ALL}")
        return None

def print_header():
    """Print a modern header for the application."""
    print(f"\n{Fore.BLUE}{'═' * 60}{Style.RESET_ALL}")
    print(f"{Fore.CYAN}   📧 BCA Paperwork Email System{Style.RESET_ALL}")
    print(f"{Fore.BLUE}{'═' * 60}{Style.RESET_ALL}")

def print_status(message, status="info"):
    """Print a status message with appropriate formatting."""
    if status == "info":
        print(f"{Fore.CYAN}➜ {message}{Style.RESET_ALL}")
    elif status == "success":
        print(f"{Fore.GREEN}✓ {message}{Style.RESET_ALL}")
    elif status == "error":
        print(f"{Fore.RED}✗ {message}{Style.RESET_ALL}")

def select_week():
    """Let user select a week from the available options."""
    sundays, last_sunday, next_sunday = get_week_dates()
    
    print(f"\n{Fore.CYAN}Select Work Week:{Style.RESET_ALL}")
    print(f"{Fore.WHITE}Available weeks:{Style.RESET_ALL}")
    for i, sunday in enumerate(sundays, 1):
        week_start = (datetime.strptime(sunday, "%A %d-%m-%Y") - timedelta(days=6)).strftime("%A %d-%m-%Y")
        print(f"{Fore.WHITE}{i}. {Fore.YELLOW}{week_start} to {sunday}{Style.RESET_ALL}")
    
    choice = input(f"\n{Fore.CYAN}Enter week number (1-{len(sundays)}):{Style.RESET_ALL} ").strip()
    try:
        week_idx = int(choice) - 1
        if not (0 <= week_idx < len(sundays)):
            print_status("Invalid selection.", "error")
            return None
        
        # Calculate selected Sunday based on the index
        if week_idx == 0:  # Next Sunday
            selected_sunday = next_sunday
        else:  # Past Sundays
            selected_sunday = last_sunday - timedelta(weeks=week_idx-1)
        
        return selected_sunday
        
    except ValueError:
        print_status("Invalid input.", "error")
        return None

def send_email(week_end_date, files, week_summary):
    """Send email with PDF attachments and week summary."""
    try:
        config = load_email_config()
        if not config:
            print_status("Email configuration not found.", "error")
            return False
        
        # Calculate week start (Monday)
        week_start = week_end_date - timedelta(days=6)
        
        # Create email message
        msg = MIMEMultipart()
        msg['From'] = config['Email']['sender_email']
        msg['To'] = config['Email']['recipient_email']
        msg['Subject'] = f"Paperwork for work week {week_end_date.strftime('%d-%m-%Y')}"
        
        # Create email body
        body = f"""Please find attached the paperwork for the week ending {week_end_date.strftime('%d-%m-%Y')}.

Week Summary:
- Total Loads: {week_summary['total_loads']}
- Total Vehicles: {week_summary['total_vehicles']}
- Date Range: {week_start.strftime('%d-%m-%Y')} to {week_end_date.strftime('%d-%m-%Y')}

Files Attached:
"""
        # Get all files from the email folder
        email_dir = os.path.join(SCRIPT_DIR, "email", week_end_date.strftime("%d-%m-%Y"))
        
        # Separate files by type
        pdf_files = [f for f in files if f.endswith('.pdf')]
        png_files = [f for f in files if f.endswith('.png')]
        
        # Count timesheet and loadsheets
        timesheet_count = sum(1 for f in pdf_files if f.startswith('timesheet_'))
        loadsheet_count = sum(1 for f in pdf_files if not f.startswith('timesheet_'))
        
        # Add PDF files to body
        if pdf_files:
            body += "\nPDF Documents:\n"
            # Add timesheet first if present
            timesheet = next((f for f in pdf_files if f.startswith('timesheet_')), None)
            if timesheet:
                body += f"- {timesheet}\n"
            # Add loadsheets
            for file in sorted(f for f in pdf_files if not f.startswith('timesheet_')):
                body += f"- {file}\n"
        
        # Add PNG files to body
        if png_files:
            body += "\nReceipts & Additional Paperwork:\n"
            for file in sorted(png_files):
                body += f"- {file}\n"
        
        # Show clean email summary
        print(f"\n{Fore.CYAN}Email Summary:{Style.RESET_ALL}")
        print(f"{Fore.WHITE}To: {config['Email']['recipient_email']}")
        print(f"{Fore.WHITE}Subject: {msg['Subject']}")
        if timesheet_count > 0:
            print(f"{Fore.GREEN}✓ Timesheet Added{Style.RESET_ALL}")
        print(f"{Fore.WHITE}Loadsheets Added: {loadsheet_count}")
        print(f"{Fore.WHITE}Receipts & Paperwork: {len(png_files)} files")
        print(f"{Fore.WHITE}Total Loads: {week_summary['total_loads']}")
        print(f"{Fore.WHITE}Total Vehicles: {week_summary['total_vehicles']}")
        
        # Ask for confirmation to send email
        confirm = input(f"\n{Fore.YELLOW}Send this email? (y/n):{Style.RESET_ALL} ").strip().lower()
        if confirm != 'y':
            print_status("Email sending cancelled.", "info")
            return False
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Attach files
        for file in files:
            file_path = os.path.join(email_dir, file)
            with open(file_path, 'rb') as f:
                # Determine file type
                if file.lower().endswith('.pdf'):
                    attachment = MIMEApplication(f.read(), _subtype='pdf')
                elif file.lower().endswith('.png'):
                    attachment = MIMEApplication(f.read(), _subtype='png')
                elif file.lower().endswith(('.jpg', '.jpeg')):
                    attachment = MIMEApplication(f.read(), _subtype='jpeg')
                elif file.lower().endswith('.gif'):
                    attachment = MIMEApplication(f.read(), _subtype='gif')
                elif file.lower().endswith('.img'):
                    attachment = MIMEApplication(f.read(), _subtype='img')
                
                attachment.add_header('Content-Disposition', 'attachment', 
                                   filename=file)
                msg.attach(attachment)
        
        # Send email
        with smtplib.SMTP(config['Email']['smtp_server'], 
                         int(config['Email']['smtp_port'])) as server:
            server.starttls()
            server.login(config['Email']['sender_email'], 
                        config['Email']['sender_password'])
            server.send_message(msg)
        
        return True
        
    except Exception as e:
        print_status(f"Failed to send email: {str(e)}", "error")
        return False

def convert_excel_to_pdf(excel_path, pdf_path):
    """Convert Excel file to PDF using LibreOffice."""
    try:
        # Convert paths to absolute paths
        excel_path = os.path.abspath(excel_path)
        pdf_path = os.path.abspath(pdf_path)
        
        logging.info(f"Converting Excel file to PDF: {excel_path}")
        
        # Check if LibreOffice is installed
        libreoffice_paths = [
            # Windows paths
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
            # Linux paths
            "/usr/bin/soffice",
            "/usr/lib/libreoffice/program/soffice",
            "/usr/lib64/libreoffice/program/soffice",
            # Try if it's in PATH
            "soffice",
        ]
        
        # Check if we're on Linux
        is_linux = os.name == 'posix'
        
        soffice_path = None
        for path in libreoffice_paths:
            if os.path.exists(path):
                soffice_path = path
                logging.info(f"Found LibreOffice at: {path}")
                break
        
        if not soffice_path:
            logging.error("LibreOffice not found. Please install LibreOffice.")
            return False
        
        # Check if source file exists
        if not os.path.exists(excel_path):
            logging.error(f"Source Excel file not found: {excel_path}")
            return False
            
        # Check if target directory exists
        pdf_dir = os.path.dirname(pdf_path)
        if not os.path.exists(pdf_dir):
            os.makedirs(pdf_dir)
            logging.info(f"Created target directory: {pdf_dir}")
        
        # Try different conversion methods based on OS
        if is_linux:
            conversion_methods = [
                # Linux methods
                [soffice_path, '--headless', '--convert-to', 'pdf', '--outdir', pdf_dir, excel_path],
                [soffice_path, '--headless', '--convert-to', 'pdf:writer_pdf_Export', '--outdir', pdf_dir, excel_path],
                [soffice_path, '--headless', '--convert-to', 'pdf', '--outdir', pdf_dir, '--infilter="MS Excel 2007 XML"', excel_path],
                # Try with full path
                [soffice_path, '--headless', '--convert-to', 'pdf', '--outdir', pdf_dir, '--infilter="Calc MS Excel 2007 XML"', excel_path]
            ]
        else:
            conversion_methods = [
                # Windows methods
                [soffice_path, '--headless', '--convert-to', 'pdf', '--outdir', pdf_dir, excel_path],
                [soffice_path, '--headless', '--convert-to', 'pdf:writer_pdf_Export', '--outdir', pdf_dir, excel_path],
                [soffice_path, '--headless', '--convert-to', 'pdf', '--outdir', pdf_dir, '--infilter="MS Excel 2007 XML"', excel_path]
            ]
        
        for method in conversion_methods:
            logging.info(f"Trying conversion method: {' '.join(method)}")
            result = subprocess.run(method, capture_output=True, text=True)
            
            if result.returncode == 0:
                # Check for the expected PDF file
                expected_pdf = os.path.splitext(os.path.basename(excel_path))[0] + '.pdf'
                expected_pdf_path = os.path.join(pdf_dir, expected_pdf)
                
                if os.path.exists(expected_pdf_path):
                    # Move the PDF to the desired location
                    shutil.move(expected_pdf_path, pdf_path)
                    logging.info(f"PDF created successfully: {pdf_path}")
                    return True
                else:
                    logging.warning(f"PDF file not found at expected location: {expected_pdf_path}")
            else:
                logging.warning(f"Conversion failed with method {' '.join(method)}")
                logging.warning(f"Error output: {result.stderr}")
        
        logging.error("All conversion methods failed")
        return False
            
    except Exception as e:
        logging.error(f"Error converting {excel_path} to PDF: {e}", exc_info=True)
        return False

def get_week_dates():
    """Get list of recent Sundays plus current/following week."""
    today = datetime.now()
    current_weekday = today.weekday()
    days_to_last_sunday = (current_weekday + 1) % 7
    last_sunday = today - timedelta(days=days_to_last_sunday)
    
    logging.info(f"Today: {today.strftime('%A %d-%m-%Y')}")
    logging.info(f"Current weekday: {current_weekday}")
    logging.info(f"Days to last Sunday: {days_to_last_sunday}")
    logging.info(f"Last Sunday: {last_sunday.strftime('%A %d-%m-%Y')}")
    
    # Show last 4 Sundays plus current/following week
    sundays = [(last_sunday - timedelta(weeks=i)).strftime("%A %d-%m-%Y") 
              for i in range(4)]
    
    # Add current/following week if we're not already showing it
    next_sunday = last_sunday + timedelta(weeks=1)
    if next_sunday.strftime("%A %d-%m-%Y") not in sundays:
        sundays.insert(0, next_sunday.strftime("%A %d-%m-%Y"))
    
    return sundays, last_sunday, next_sunday

def check_paperwork_files(selected_sunday):
    """Check for loadsheets and timesheet, convert to PDF, and organize in email folder."""
    try:
        # Calculate week start (Monday)
        week_start = selected_sunday - timedelta(days=6)
        week_end = selected_sunday
        
        print_status(f"Processing week: {week_start.strftime('%d-%m-%Y')} to {week_end.strftime('%d-%m-%Y')}")
        
        # Create email folder
        email_dir = os.path.join(SCRIPT_DIR, "email")
        week_email_dir = os.path.join(email_dir, week_end.strftime("%d-%m-%Y"))
        os.makedirs(week_email_dir, exist_ok=True)
        
        # Check timesheet directory first
        timesheets_dir = os.path.join(SCRIPT_DIR, "timesheets")
        timesheet_folder = os.path.join(timesheets_dir, week_end.strftime("%Y%m%d"))
        timesheet_file = os.path.join(timesheet_folder, f"timesheet_{week_end.strftime('%Y%m%d')}.xlsx")
        
        if not os.path.exists(timesheet_file):
            print_status("Timesheet not found.", "error")
            return False
        
        print_status("Converting timesheet to PDF...")
        # Convert timesheet to PDF
        timesheet_pdf = f"timesheet_{week_end.strftime('%Y%m%d')}.pdf"
        timesheet_pdf_path = os.path.join(week_email_dir, timesheet_pdf)
        
        if convert_excel_to_pdf(timesheet_file, timesheet_pdf_path):
            print_status("Timesheet converted successfully", "success")
        else:
            print_status("Failed to convert timesheet", "error")
            return False
        
        # Get week summary from timesheet
        week_summary = get_week_summary(week_end, timesheet_file)
        if not week_summary:
            print_status("Failed to get week summary", "error")
            return False
        
        # Check loadsheets directory only if there are loads
        if week_summary['total_loads'] > 0:
            # Check loadsheets directory
            loadsheets_dir = os.path.join(SCRIPT_DIR, "loadsheets")
            week_folder = os.path.join(loadsheets_dir, week_end.strftime("%d-%m-%Y"))
            
            if not os.path.exists(week_folder):
                print_status("Loadsheets folder not found.", "error")
                return False
            
            # Get list of loadsheets
            loadsheets = [f for f in os.listdir(week_folder) if f.endswith('.xlsx')]
            print_status(f"Converting {len(loadsheets)} loadsheets to PDF...")
            
            # Convert loadsheets to PDF
            converted_loadsheets = []
            for loadsheet in sorted(loadsheets):
                excel_path = os.path.join(week_folder, loadsheet)
                pdf_name = os.path.splitext(loadsheet)[0] + '.pdf'
                pdf_path = os.path.join(week_email_dir, pdf_name)
                
                if convert_excel_to_pdf(excel_path, pdf_path):
                    converted_loadsheets.append(pdf_name)
            
            print_status(f"Converted {len(converted_loadsheets)} loadsheets", "success")
        
        # Get list of PDF and PNG files
        pdf_files = [f for f in os.listdir(week_email_dir) if f.endswith('.pdf')]
        png_files = [f for f in os.listdir(week_email_dir) if f.endswith('.png')]
        
        # Send email
        if send_email(week_end, pdf_files + png_files, week_summary):
            print_status("Email sent successfully!", "success")
            return True
        else:
            print_status("Failed to send email", "error")
            return False
        
    except Exception as e:
        print_status(f"Error: {str(e)}", "error")
        return False

def main():
    """Main function to run the email preparation process."""
    print_header()
    
    try:
        # Test SMTP connection first
        if not test_smtp_connection():
            print_status("Please check your email configuration and try again.", "error")
            return
        
        # Select week
        selected_sunday = select_week()
        if not selected_sunday:
            return
        
        # Check paperwork files
        if check_paperwork_files(selected_sunday):
            print_status("Paperwork processing completed successfully!", "success")
        else:
            print_status("Failed to process paperwork.", "error")
            
    except KeyboardInterrupt:
        print_status("Operation interrupted.", "info")
    except Exception as e:
        print_status(f"An unexpected error occurred: {str(e)}", "error")
    finally:
        print(f"\n{Fore.GREEN}Program terminated.{Style.RESET_ALL}")

if __name__ == "__main__":
    main() 