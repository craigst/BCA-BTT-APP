#!/usr/bin/env python3
"""
Paperwork Module

This module provides functions to generate:
  1. A loadsheet for a given load number
  2. A timesheet for a given week‐ending date
  3. All loadsheets and timesheet for a selected work week
  4. Test loadsheet generation with debug information
"""

import os
import random
import configparser
import logging
from datetime import datetime, timedelta, date
import psycopg2
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
import colorama
from colorama import Fore, Back, Style
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker, AbsoluteAnchor, XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU, EMU_to_pixels
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

# Initialize colorama for cross-platform terminal colors
colorama.init(autoreset=True)

# -------- Logging Setup --------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LOG_DIR = os.path.join(SCRIPT_DIR, "logs")
SQL_DIR = os.path.join(SCRIPT_DIR, "sql")

# Ensure directories exist
os.makedirs(LOG_DIR, exist_ok=True)
os.makedirs(SQL_DIR, exist_ok=True)

# Setup logging configuration
LOG_FILE = os.path.join(LOG_DIR, f"paperwork_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE)
    ]
)

def print_header():
    """Print a modern header for the application."""
    print(f"\n{Fore.BLUE}{'═' * 60}{Style.RESET_ALL}")
    print(f"{Fore.CYAN}   📄 BCA Paperwork System{Style.RESET_ALL}")
    print(f"{Fore.BLUE}{'═' * 60}{Style.RESET_ALL}")

def print_status(message, status="info"):
    """Print a status message with appropriate formatting."""
    if status == "info":
        print(f"{Fore.CYAN}➜ {message}{Style.RESET_ALL}")
    elif status == "success":
        print(f"{Fore.GREEN}✓ {message}{Style.RESET_ALL}")
    elif status == "error":
        print(f"{Fore.RED}✗ {message}{Style.RESET_ALL}")

def print_menu():
    """Print a modern menu interface."""
    print(f"\n{Fore.CYAN}Available Options:{Style.RESET_ALL}")
    print(f"{Fore.WHITE}1. {Fore.YELLOW}Create All Paperwork{Style.RESET_ALL}")
    print(f"{Fore.WHITE}2. {Fore.YELLOW}Create Single Loadsheet{Style.RESET_ALL}")
    print(f"{Fore.WHITE}3. {Fore.YELLOW}Create Timesheet{Style.RESET_ALL}")
    print(f"{Fore.WHITE}4. {Fore.YELLOW}Toggle Auto Signature{Style.RESET_ALL}")
    print(f"{Fore.WHITE}5. {Fore.YELLOW}Exit{Style.RESET_ALL}")

class SignatureConfig:
    """Configuration for signature placement and appearance."""
    def __init__(self):
        # Base scale (1.0 = original size)
        self.scale = 1.2  # 20% larger than original
        
        # Fine-tuning offsets (in pixels)
        self.sig1_offset_x = 0  # Adjust left/right (0 = centered)
        self.sig1_offset_y = -30  # Adjust up/down (negative moves up)
        self.sig2_offset_x = 0  # Adjust left/right (0 = centered)
        self.sig2_offset_y = -30  # Adjust up/down (negative moves up)
        
        # Random movement ranges (in pixels)
        self.random_x_range = (0.1, 0.3)  # Min and max random X movement
        self.random_y_range = (0.1, 0.2)  # Min and max random Y movement
        
        # Random rotation range (in degrees)
        self.random_rotation_range = (-3, 3)  # Min and max random rotation
        
        # Base cell positions (can be overridden)
        self.sig1_cell = 'C44'
        self.sig2_cell = 'H44'
        
        # Cell dimensions (approximate)
        self.cell_width_px = 8  # Approximate cell width in pixels
        self.cell_height_px = 15  # Approximate cell height in pixels
        
        # Allow cell overlap
        self.allow_overlap = True
    
    def get_random_offset(self):
        """Get random X and Y offsets within configured ranges."""
        import random
        x_offset = random.uniform(self.random_x_range[0], self.random_x_range[1])
        y_offset = random.uniform(self.random_y_range[0], self.random_y_range[1])
        rotation = random.uniform(self.random_rotation_range[0], self.random_rotation_range[1])
        return x_offset, y_offset, rotation
    
    def get_sig1_position(self):
        """Get final position for signature 1 with all offsets applied."""
        x_offset, y_offset, rotation = self.get_random_offset()
        return {
            'cell': self.sig1_cell,
            'offset_x': self.sig1_offset_x + x_offset,
            'offset_y': self.sig1_offset_y + y_offset,
            'rotation': rotation,
            'allow_overlap': self.allow_overlap,
            'cell_width': self.cell_width_px,
            'cell_height': self.cell_height_px
        }
    
    def get_sig2_position(self):
        """Get final position for signature 2 with all offsets applied."""
        x_offset, y_offset, rotation = self.get_random_offset()
        return {
            'cell': self.sig2_cell,
            'offset_x': self.sig2_offset_x + x_offset,
            'offset_y': self.sig2_offset_y + y_offset,
            'rotation': rotation,
            'allow_overlap': self.allow_overlap,
            'cell_width': self.cell_width_px,
            'cell_height': self.cell_height_px
        }

class PaperworkManager:
    def __init__(self):
        """Initialize the paperwork manager with PostgreSQL configuration."""
        self.pg_config = self.load_pg_config()
        self.config_file = os.path.join(SCRIPT_DIR, "config.ini")
        self.auto_signature = self.load_auto_signature_config()
        
    def load_pg_config(self):
        """Load PostgreSQL configuration from sql.ini file."""
        config = configparser.ConfigParser()
        config_path = os.path.join(SQL_DIR, "sql.ini")
        
        if not os.path.exists(config_path):
            logging.error(f"PostgreSQL configuration file not found at {config_path}")
            return None
            
        try:
            config.read(config_path)
            return {
                'host': config['SQL']['PG_HOST'],
                'port': config['SQL']['PG_PORT'],
                'database': config['SQL']['PG_DATABASE'],
                'user': config['SQL']['PG_USERNAME'],
                'password': config['SQL']['PG_PASSWORD']
            }
        except Exception as e:
            logging.error(f"Error loading PostgreSQL configuration: {e}")
            return None

    def load_auto_signature_config(self):
        """Load auto signature setting from config file."""
        config = configparser.ConfigParser()
        try:
            if os.path.exists(self.config_file):
                config.read(self.config_file)
                return config.getboolean('Settings', 'auto_signature', fallback=True)
        except Exception as e:
            logging.error(f"Error loading auto signature config: {e}")
        return True  # Default to True if config file doesn't exist or has error
        
    def save_auto_signature_config(self):
        """Save auto signature setting to config file."""
        config = configparser.ConfigParser()
        config['Settings'] = {
            'auto_signature': str(self.auto_signature).lower()
        }
        try:
            with open(self.config_file, 'w') as f:
                config.write(f)
        except Exception as e:
            logging.error(f"Error saving auto signature config: {e}")

    def get_week_dates(self):
        """Get list of recent Sundays plus current/following week."""
        today = datetime.now()
        current_weekday = today.weekday()
        days_to_last_sunday = (current_weekday + 1) % 7
        last_sunday = today - timedelta(days=days_to_last_sunday)
        
        logging.info(f"Today: {today.strftime('%A %d-%m-%Y')}")
        logging.info(f"Current weekday: {current_weekday}")
        logging.info(f"Days to last Sunday: {days_to_last_sunday}")
        logging.info(f"Last Sunday: {last_sunday.strftime('%A %d-%m-%Y')}")
        
        # Show last 4 Sundays plus current/following week
        sundays = [(last_sunday - timedelta(weeks=i)).strftime("%A %d-%m-%Y") 
                  for i in range(4)]
        
        # Add current/following week if we're not already showing it
        next_sunday = last_sunday + timedelta(weeks=1)
        if next_sunday.strftime("%A %d-%m-%Y") not in sundays:
            sundays.insert(0, next_sunday.strftime("%A %d-%m-%Y"))
        
        return sundays, last_sunday, next_sunday

    def select_week(self):
        """Let user select a week from the available options."""
        sundays, last_sunday, next_sunday = self.get_week_dates()
        
        print(f"\n{Fore.CYAN}Select week end date (Sunday):{Style.RESET_ALL}")
        for i, sunday in enumerate(sundays, 1):
            print(f"{Fore.WHITE}{i}. {Fore.YELLOW}{sunday}{Style.RESET_ALL}")
        
        choice = input(f"\n{Fore.CYAN}Enter week number (1-{len(sundays)}):{Style.RESET_ALL} ").strip()
        try:
            week_idx = int(choice) - 1
            if not (0 <= week_idx < len(sundays)):
                print(f"{Fore.RED}Invalid selection.{Style.RESET_ALL}")
                return None
            
            # Calculate selected Sunday based on the index
            if week_idx == 0:  # Next Sunday
                selected_sunday = next_sunday
            else:  # Past Sundays
                selected_sunday = last_sunday - timedelta(weeks=week_idx-1)
            
            return selected_sunday
            
        except ValueError:
            print(f"{Fore.RED}Invalid input.{Style.RESET_ALL}")
            return None

    def get_loads_for_week(self, selected_sunday):
        """Get all loads for the selected week."""
        week_start = selected_sunday - timedelta(days=6)  # Monday
        start_date_str = week_start.strftime("%Y%m%d")
        end_date_str = selected_sunday.strftime("%Y%m%d")
        
        try:
            conn = psycopg2.connect(**self.pg_config)
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT DISTINCT dwvload
                FROM public.dwvveh
                WHERE dwvexpdat BETWEEN %s AND %s
                AND dwvload IS NOT NULL
                ORDER BY dwvload
            """, (start_date_str, end_date_str))
            
            loads = cursor.fetchall()
            cursor.close()
            conn.close()
            
            return loads
            
        except Exception as e:
            logging.error(f"Error getting loads for week: {e}")
            return []

    def get_load_info(self, load_number):
        """Get detailed information for a specific load."""
        try:
            conn = psycopg2.connect(**self.pg_config)
            cursor = conn.cursor()
            
            # Get collections with vehicle details
            cursor.execute("""
                SELECT 
                    j.dwjtype,
                    j.dwjcust,
                    j.dwjname,
                    j.dwjdate,
                    j.dwjadrcod,
                    j.dwjpostco,
                    j.dwjvehs,
                    v.dwvvehref,
                    v.dwvmoddes,
                    COALESCE(e.sparekeys, 'Y') as sparekeys,
                    COALESCE(e.extra, 'Y') as extra,
                    COALESCE(e.carnotes, '') as carnotes,
                    v.dwvcolcod,
                    v.dwvdelcod
                FROM public.dwjjob j
                LEFT JOIN public.dwvveh v ON j.dwjload = v.dwvload AND j.dwjadrcod = v.dwvcolcod
                LEFT JOIN public.extracarinfo e ON v.dwvkey = e.idkey
                WHERE j.dwjload = %s
                AND j.dwjtype = 'C'
                ORDER BY j.dwjdate, j.dwjcust
            """, (load_number,))
            
            collections = cursor.fetchall()
            
            # Get deliveries with vehicle details
            cursor.execute("""
                SELECT 
                    j.dwjtype,
                    j.dwjcust,
                    j.dwjname,
                    j.dwjdate,
                    j.dwjadrcod,
                    j.dwjpostco,
                    j.dwjvehs,
                    v.dwvvehref,
                    v.dwvmoddes,
                    COALESCE(e.sparekeys, 'Y') as sparekeys,
                    COALESCE(e.extra, 'Y') as extra,
                    COALESCE(e.carnotes, '') as carnotes,
                    v.dwvcolcod,
                    v.dwvdelcod
                FROM public.dwjjob j
                LEFT JOIN public.dwvveh v ON j.dwjload = v.dwvload AND j.dwjadrcod = v.dwvdelcod
                LEFT JOIN public.extracarinfo e ON v.dwvkey = e.idkey
                WHERE j.dwjload = %s
                AND j.dwjtype = 'D'
                ORDER BY j.dwjdate, j.dwjcust
            """, (load_number,))
            
            deliveries = cursor.fetchall()
            
            if not collections and not deliveries:
                print(f"{Fore.YELLOW}No details found for this load.{Style.RESET_ALL}")
                return None
            
            # Format collections with postcodes and vehicle details
            formatted_collections = []
            collection_vehicles = []
            for collection in collections:
                if collection[2] and collection[5]:  # Check if name and postcode exist
                    formatted_collections.append(f"{collection[2]} - {collection[5]}")
                elif collection[2]:  # If only name exists
                    formatted_collections.append(collection[2])
                
                # Add vehicle details if available
                if collection[7] and collection[8]:  # Check if vehicle reference and model exist
                    collection_vehicles.append(f"{collection[7]} - {collection[8]}")
            
            # Format deliveries with postcodes and vehicle details
            formatted_deliveries = []
            delivery_vehicles = []
            for delivery in deliveries:
                if delivery[2] and delivery[5]:  # Check if name and postcode exist
                    formatted_deliveries.append(f"{delivery[2]} - {delivery[5]}")
                elif delivery[2]:  # If only name exists
                    formatted_deliveries.append(delivery[2])
                
                # Add vehicle details if available
                if delivery[7] and delivery[8]:  # Check if vehicle reference and model exist
                    delivery_vehicles.append(f"{delivery[7]} - {delivery[8]}")
            
            # Get the first collection date and contractor
            first_collection_date = collections[0][3] if collections else None
            contractor = collections[0][1] if collections else None
            
            # Format the data for the loadsheet
            load_info = (
                first_collection_date,  # date
                load_number,  # load number
                '\n'.join(formatted_collections),  # collections
                '\n'.join(formatted_deliveries),  # deliveries
                len(collection_vehicles),  # collection cars count
                len(delivery_vehicles),  # delivery cars count
                contractor  # contractor
            )
            
            # Format vehicle data for the loadsheet
            formatted_vehicles = []
            for collection in collections:
                if collection[7] and collection[8]:  # Check if vehicle reference and model exist
                    formatted_vehicles.append((
                        str(collection[7] or ''),  # registration
                        str(collection[8] or ''),  # model
                        'N',  # offloaded (default)
                        'Y',  # documents (default)
                        str(collection[7] or ''),  # key (using registration as key)
                        str(collection[9] or 'Y'),  # spare keys
                        str(collection[11] or '')   # notes
                    ))
            
            cursor.close()
            conn.close()
            
            return {
                'load_info': load_info,
                'vehicles': formatted_vehicles
            }
            
        except Exception as e:
            logging.error(f"Error getting load info: {e}", exc_info=True)
            print(f"{Fore.RED}Error getting load info: {e}{Style.RESET_ALL}")
            return None

    def show_load_summary(self, load_data):
        """Show a summary of the load information."""
        if not load_data:
            print(f"{Fore.RED}No data available for this load.{Style.RESET_ALL}")
            return False
            
        load_info, vehicles = load_data
        
        print(f"\n{Fore.CYAN}Load {load_info[1]} Details:{Style.RESET_ALL}")
        
        # Show collections
        if load_info[2]:
            print(f"\n{Fore.YELLOW}Collections:{Style.RESET_ALL}")
            print(f"{Fore.WHITE}{'Type':<8} | {'Customer':<10} | {'Location':<30} | {'Cars':<5}{Style.RESET_ALL}")
            print("-" * 60)
            # Create a set of unique collections based on name and postcode
            unique_collections = {}
            for collection in load_info[2].split('\n'):
                if collection:
                    if collection not in unique_collections:
                        unique_collections[collection] = 1
                    else:
                        unique_collections[collection] += 1
            
            # Display unique collections
            for location, count in sorted(unique_collections.items()):
                print(f"{Fore.WHITE}{'C':<8} | {Fore.YELLOW}{location:<40} | {Fore.YELLOW}{count:<5}{Style.RESET_ALL}")
        
        # Show deliveries
        if load_info[3]:
            print(f"\n{Fore.YELLOW}Deliveries:{Style.RESET_ALL}")
            print(f"{Fore.WHITE}{'Type':<8} | {'Customer':<10} | {'Location':<30} | {'Cars':<5}{Style.RESET_ALL}")
            print("-" * 60)
            # Create a set of unique deliveries based on name and postcode
            unique_deliveries = {}
            for delivery in load_info[3].split('\n'):
                if delivery:
                    if delivery not in unique_deliveries:
                        unique_deliveries[delivery] = 1
                    else:
                        unique_deliveries[delivery] += 1
            
            # Display unique deliveries
            for location, count in sorted(unique_deliveries.items()):
                print(f"{Fore.WHITE}{'D':<8} | {Fore.YELLOW}{location:<40} | {Fore.YELLOW}{count:<5}{Style.RESET_ALL}")
        
        # Show vehicles
        if vehicles:
            print(f"\n{Fore.YELLOW}Vehicles:{Style.RESET_ALL}")
            print(f"{Fore.WHITE}{'Reg':<10} | {'Model':<30} | {'Offloaded':<10} | {'Documents':<10} | {'Spare Keys':<10} | {'Notes'}{Style.RESET_ALL}")
            print("-" * 100)
            for vehicle in vehicles:
                try:
                    if not isinstance(vehicle, (list, tuple)) or len(vehicle) < 7:
                        print(f"{Fore.RED}Warning: Invalid vehicle data format, skipping.{Style.RESET_ALL}")
                        continue
                        
                    print(f"{Fore.WHITE}{str(vehicle[0] or ''):<10} | {Fore.YELLOW}{str(vehicle[1] or ''):<30} | {Fore.YELLOW}{str(vehicle[2] or ''):<10} | {Fore.YELLOW}{str(vehicle[3] or ''):<10} | {Fore.YELLOW}{str(vehicle[5] or ''):<10} | {Fore.YELLOW}{str(vehicle[6] or '')}{Style.RESET_ALL}")
                except Exception as e:
                    print(f"{Fore.RED}Warning: Error displaying vehicle: {e}{Style.RESET_ALL}")
                    continue
        
        return True

    def create_loadsheet(self, load_number):
        """Create a loadsheet for the specified load."""
        wb = None
        try:
            # Create loadsheets directory if it doesn't exist
            loadsheets_dir = os.path.join(SCRIPT_DIR, "loadsheets")
            os.makedirs(loadsheets_dir, exist_ok=True)
            logging.info(f"Created/verified loadsheets directory: {loadsheets_dir}")
            
            # Get load info directly from database
            try:
                conn = psycopg2.connect(**self.pg_config)
                cursor = conn.cursor()
                logging.info("Database connection successful")
                print(f"{Fore.GREEN}Database connection successful.{Style.RESET_ALL}")
            except Exception as e:
                logging.error(f"Database connection failed: {e}")
                print(f"{Fore.RED}Database connection failed: {e}{Style.RESET_ALL}")
                return False
            
            try:
                # Get collections
                cursor.execute("""
                    SELECT 
                        j.dwjtype,
                        j.dwjcust,
                        j.dwjname,
                        j.dwjdate,
                        j.dwjadrcod,
                        j.dwjpostco,
                        j.dwjvehs,
                        v.dwvvehref,
                        v.dwvmoddes,
                        COALESCE(e.sparekeys, 'Y') as sparekeys,
                        COALESCE(e.extra, 'Y') as extra,
                        COALESCE(e.carnotes, '') as carnotes,
                        v.dwvcolcod,
                        v.dwvdelcod
                    FROM public.dwjjob j
                    LEFT JOIN public.dwvveh v ON j.dwjload = v.dwvload AND j.dwjadrcod = v.dwvcolcod
                    LEFT JOIN public.extracarinfo e ON v.dwvkey = e.idkey
                    WHERE j.dwjload = %s
                    AND j.dwjtype = 'C'
                    ORDER BY j.dwjdate, j.dwjcust
                """, (load_number,))
                
                collections = cursor.fetchall()
                logging.info(f"Found {len(collections)} collections for load {load_number}")
                
                # Get deliveries
                cursor.execute("""
                    SELECT 
                        j.dwjtype,
                        j.dwjcust,
                        j.dwjname,
                        j.dwjdate,
                        j.dwjadrcod,
                        j.dwjpostco,
                        j.dwjvehs,
                        v.dwvvehref,
                        v.dwvmoddes,
                        COALESCE(e.sparekeys, 'Y') as sparekeys,
                        COALESCE(e.extra, 'Y') as extra,
                        COALESCE(e.carnotes, '') as carnotes,
                        v.dwvcolcod,
                        v.dwvdelcod
                    FROM public.dwjjob j
                    LEFT JOIN public.dwvveh v ON j.dwjload = v.dwvload AND j.dwjadrcod = v.dwvdelcod
                    LEFT JOIN public.extracarinfo e ON v.dwvkey = e.idkey
                    WHERE j.dwjload = %s
                    AND j.dwjtype = 'D'
                    ORDER BY j.dwjdate, j.dwjcust
                """, (load_number,))
                
                deliveries = cursor.fetchall()
                logging.info(f"Found {len(deliveries)} deliveries for load {load_number}")
                
                # Get vehicles
                cursor.execute("""
                    SELECT DISTINCT
                        v.dwvvehref,
                        v.dwvmoddes,
                        v.dwvcolcod,
                        v.dwvdelcod,
                        COALESCE(e.sparekeys, 'Y') as sparekeys,
                        COALESCE(e.extra, 'Y') as extra,
                        COALESCE(e.carnotes, '') as carnotes
                    FROM public.dwvveh v
                    LEFT JOIN public.extracarinfo e ON v.dwvkey = e.idkey
                    WHERE v.dwvload = %s
                    ORDER BY v.dwvvehref
                """, (load_number,))
                
                vehicles = cursor.fetchall()
                logging.info(f"Found {len(vehicles)} vehicles for load {load_number}")
                
                if not collections and not deliveries:
                    logging.warning(f"No details found for load {load_number}")
                    print(f"{Fore.YELLOW}No details found for this load.{Style.RESET_ALL}")
                    return False
                
                # Get the date from the first collection or delivery
                date_str = None
                if collections:
                    date_str = str(collections[0][3])
                elif deliveries:
                    date_str = str(deliveries[0][3])
                
                if not date_str:
                    logging.error(f"No date found for load {load_number}")
                    print(f"{Fore.RED}No date found for load {load_number}{Style.RESET_ALL}")
                    return False
                
                # Convert date to datetime
                load_date = datetime.strptime(date_str, '%Y%m%d')
                logging.info(f"Load date: {load_date.strftime('%Y-%m-%d')}")
                
                # Create week folder with Sunday's date
                week_end = load_date + timedelta(days=(6 - load_date.weekday()))
                week_folder = os.path.join(loadsheets_dir, week_end.strftime("%d-%m-%Y"))
                os.makedirs(week_folder, exist_ok=True)
                logging.info(f"Created/verified week folder: {week_folder}")
                
                # Get town name from first collection or delivery
                town_name = ""
                if collections:
                    town_name = collections[0][2] or ""  # dwjname from first collection
                elif deliveries:
                    town_name = deliveries[0][2] or ""  # dwjname from first delivery
                logging.info(f"Town name: {town_name}")
                
                # Set up the output file path
                output_file = os.path.join(week_folder, f"{load_number}_{load_date.strftime('%d-%m-%Y')}_{town_name}.xlsx")
                logging.info(f"Output file path: {output_file}")
                
                # Copy template to output file
                template_path = os.path.join(SCRIPT_DIR, "templates", "loadsheet.xlsx")
                if not os.path.exists(template_path):
                    logging.error(f"Template file not found at {template_path}")
                    print(f"{Fore.RED}Template file not found at {template_path}{Style.RESET_ALL}")
                    return False
                
                # Remove any existing lock files
                lock_file = os.path.join(os.path.dirname(template_path), ".~lock.loadsheet.xlsx#")
                if os.path.exists(lock_file):
                    try:
                        os.remove(lock_file)
                        logging.info(f"Removed lock file: {lock_file}")
                    except Exception as e:
                        logging.warning(f"Could not remove lock file: {e}")
                
                import shutil
                shutil.copy2(template_path, output_file)
                logging.info(f"Copied template to output file")
                
                # Load the workbook without data_only=True
                wb = load_workbook(output_file)
                
                # Verify the worksheet exists
                if "Loadsheet" not in wb.sheetnames:
                    logging.error("Worksheet 'Loadsheet' not found in template")
                    print(f"{Fore.RED}Error: Worksheet 'Loadsheet' not found in template{Style.RESET_ALL}")
                    return False
                
                ws = wb["Loadsheet"]
                logging.info("Loaded workbook and worksheet")
                
                def safe_cell_write(cell_ref, value):
                    """Safely write to a cell."""
                    try:
                        # Convert value to string and capitalize if it's not None
                        cell_value = str(value).upper() if value is not None else ''
                        ws[cell_ref] = cell_value
                        logging.info(f"Successfully wrote value '{cell_value}' to cell {cell_ref}")
                    except Exception as e:
                        logging.error(f"Error writing to cell {cell_ref}: {e}")
                
                # Update header information
                if collections:
                    # Get the first collection date
                    date_str = str(collections[0][3])
                    try:
                        if len(date_str) == 8:  # Ensure date string is in YYYYMMDD format
                            date_obj = datetime.strptime(date_str, '%Y%m%d')
                            formatted_date = date_obj.strftime('%d/%m/%Y')
                            # Update collection date in header
                            safe_cell_write('C6', formatted_date)  # Collection date in header
                            logging.info(f"Wrote collection date {formatted_date} to C6")
                    except ValueError as e:
                        logging.warning(f"Error parsing collection date ({date_str}): {e}")
                        print(f"{Fore.YELLOW}Warning: Error parsing collection date ({date_str}): {e}{Style.RESET_ALL}")
                
                if deliveries:
                    # Get the first delivery date
                    date_str = str(deliveries[0][3])
                    try:
                        if len(date_str) == 8:  # Ensure date string is in YYYYMMDD format
                            date_obj = datetime.strptime(date_str, '%Y%m%d')
                            formatted_date = date_obj.strftime('%d/%m/%Y')
                            # Update delivery date in signature section
                            safe_cell_write('H46', formatted_date)  # Delivery date in signature section
                            logging.info(f"Wrote delivery date {formatted_date} to H46")
                    except ValueError as e:
                        logging.warning(f"Error parsing delivery date ({date_str}): {e}")
                        print(f"{Fore.YELLOW}Warning: Error parsing delivery date ({date_str}): {e}{Style.RESET_ALL}")
                
                # Update collection date in signature section
                if collections:
                    date_str = str(collections[0][3])
                    try:
                        if len(date_str) == 8:  # Ensure date string is in YYYYMMDD format
                            date_obj = datetime.strptime(date_str, '%Y%m%d')
                            formatted_date = date_obj.strftime('%d/%m/%Y')
                            safe_cell_write('C46', formatted_date)  # Collection date in signature section
                            logging.info(f"Wrote collection date {formatted_date} to C46")
                    except ValueError as e:
                        logging.warning(f"Error parsing collection date ({date_str}): {e}")
                        print(f"{Fore.YELLOW}Warning: Error parsing collection date ({date_str}): {e}{Style.RESET_ALL}")
                
                safe_cell_write('G6', str(load_number))  # Load Number
                safe_cell_write('I6', str(load_number))  # Job ID (using load number)
                logging.info(f"Wrote load number {load_number} to G6 and I6")
                
                # Update collection and delivery locations
                if collections:
                    # Create a set of unique collection locations
                    unique_collections = {}
                    for collection in collections:
                        location = f"{collection[2]} - {collection[5]}" if collection[2] and collection[5] else collection[2] or ''
                        if location not in unique_collections:
                            unique_collections[location] = 1
                        else:
                            unique_collections[location] += 1
                    
                    # Join all locations with newlines
                    collection_text = '\n'.join(unique_collections.keys())
                    safe_cell_write('B9', collection_text)
                    logging.info(f"Wrote collection locations to B9: {collection_text}")
                
                if deliveries:
                    # Create a set of unique delivery locations
                    unique_deliveries = {}
                    for delivery in deliveries:
                        location = f"{delivery[2]} - {delivery[5]}" if delivery[2] and delivery[5] else delivery[2] or ''
                        if location not in unique_deliveries:
                            unique_deliveries[location] = 1
                        else:
                            unique_deliveries[location] += 1
                    
                    # Join all locations with newlines
                    delivery_text = '\n'.join(unique_deliveries.keys())
                    safe_cell_write('F9', delivery_text)
                    logging.info(f"Wrote delivery locations to F9: {delivery_text}")
                
                # Format vehicle data for summary
                formatted_vehicles = []
                for vehicle in vehicles:
                    formatted_vehicles.append((
                        str(vehicle[0] or ''),  # registration
                        str(vehicle[1] or ''),  # model
                        'N',  # offloaded (default)
                        'Y',  # documents (default)
                        str(vehicle[4] or 'Y'),  # spare keys
                        str(vehicle[5] or 'Y'),  # extra (documents)
                        str(vehicle[6] or '')   # notes
                    ))
                logging.info(f"Formatted {len(formatted_vehicles)} vehicles")
                
                # Update vehicle information
                for i, vehicle in enumerate(formatted_vehicles[:8]):  # Handle up to 8 vehicles
                    base_row = 11 + (i * 4)  # Starting from row 11, increment by 4 for each car
                    logging.info(f"Writing vehicle {i+1} to rows {base_row}-{base_row+2}")
                    
                    # Car details - swapped registration and make & model, and capitalize all data
                    safe_cell_write(f'B{base_row}', str(vehicle[1] or '').upper())  # Make & Model
                    safe_cell_write(f'B{base_row + 2}', str(vehicle[0] or '').upper())  # Registration
                    safe_cell_write(f'E{base_row - 1}', 'N')  # Offloaded (default)
                    safe_cell_write(f'G{base_row - 1}', 'Y')  # Documents (default)
                    safe_cell_write(f'I{base_row - 1}', str(vehicle[4] or 'Y').upper())  # Spare Keys
                    if vehicle[6]:  # Notes
                        safe_cell_write(f'C{base_row}', str(vehicle[6]).upper())
                
                # Generate and add load summary message
                summary_message = self.generate_load_summary(formatted_vehicles)
                safe_cell_write('C39', summary_message.upper())  # Capitalize summary message
                logging.info(f"Wrote summary message to C39: {summary_message}")
                
                # Add signatures if enabled
                if self.auto_signature:
                    self.add_signatures(ws)
                    logging.info("Added signatures to worksheet")
                
                # Save the workbook
                wb.save(output_file)
                logging.info(f"Saved workbook to {output_file}")
                
                # Verify the file exists and has content
                if os.path.exists(output_file) and os.path.getsize(output_file) > 0:
                    print(f"{Fore.GREEN}Loadsheet created successfully at {output_file}{Style.RESET_ALL}")
                    return True
                else:
                    logging.error("File was not created or is empty")
                    print(f"{Fore.RED}Error: File was not created or is empty{Style.RESET_ALL}")
                    return False
                
            except Exception as e:
                logging.error(f"Error during query execution: {e}", exc_info=True)
                print(f"{Fore.RED}Error during query execution: {e}{Style.RESET_ALL}")
                return False
            finally:
                if 'cursor' in locals():
                    cursor.close()
                if 'conn' in locals():
                    conn.close()
                if wb is not None:
                    try:
                        wb.close()
                    except:
                        pass
            
        except Exception as e:
            logging.error(f"Error in create_loadsheet: {e}", exc_info=True)
            print(f"{Fore.RED}Error creating loadsheet: {e}{Style.RESET_ALL}")
            return False

    def create_timesheet(self, selected_sunday):
        """Create a timesheet for the selected week."""
        try:
            # Create directory for timesheets if it doesn't exist
            week_folder = os.path.join(SCRIPT_DIR, "timesheets", selected_sunday.strftime("%Y%m%d"))
            os.makedirs(week_folder, exist_ok=True)
            
            # Set up output file path
            output_file = os.path.join(week_folder, f"timesheet_{selected_sunday.strftime('%Y%m%d')}.xlsx")
            template_file = os.path.join(SCRIPT_DIR, "templates", "timesheet.xlsx")
            
            # Copy template to output file
            import shutil
            shutil.copy2(template_file, output_file)
            
            # Load the workbook
            workbook = load_workbook(output_file)
            ws = workbook["Timesheet"]
            
            # Helper functions
            def safe_cell_write(cell_ref, value):
                """Safely write a value to a cell, with error handling."""
                try:
                    # Convert value to string and capitalize if it's not None
                    cell_value = str(value).upper() if value is not None else ''
                    ws[cell_ref] = cell_value
                    logging.info(f"Successfully wrote value '{cell_value}' to cell {cell_ref}")
                except Exception as e:
                    logging.error(f"Error writing to cell {cell_ref}: {e}")
            
            def format_time(time_str):
                """Format time string for Excel."""
                return time_str.strftime("%H:%M") if time_str else ""
            
            def format_total_hours(hours):
                """Format total hours for Excel."""
                return f"{float(hours):.2f}" if hours else "0.00"
            
            # Write week end date to cell E5
            week_end_date = selected_sunday.strftime("%A %d/%m/%Y").upper()
            safe_cell_write('E5', week_end_date)
            logging.info(f"Wrote week end date to E5: {week_end_date}")
            
            # Calculate week start date
            week_start = selected_sunday - timedelta(days=6)  # Monday
            start_date_str = week_start.strftime("%Y%m%d")
            end_date_str = selected_sunday.strftime("%Y%m%d")
            
            # Fetch loads and hours for the week
            logging.info(f"Fetching data for week ending {selected_sunday.strftime('%Y-%m-%d')}")
            try:
                conn = psycopg2.connect(**self.pg_config)
                cursor = conn.cursor()
                print(f"{Fore.GREEN}Database connection successful.{Style.RESET_ALL}")
            except Exception as e:
                print(f"{Fore.RED}Database connection failed: {e}{Style.RESET_ALL}")
                return False
            
            try:
                # Get loads for the week with all required information
                cursor.execute("""
                    WITH load_dates AS (
                        SELECT 
                            j.dwjload,
                            MAX(j.dwjdate) as load_date,
                            j.dwjcust as contractor,
                            STRING_AGG(DISTINCT CASE WHEN j.dwjtype = 'C' THEN j.dwjtown ELSE NULL END, ' | ') as collections,
                            STRING_AGG(DISTINCT CASE WHEN j.dwjtype = 'D' THEN j.dwjtown ELSE NULL END, ' | ') as deliveries,
                            COUNT(DISTINCT v.dwvvehref) as total_cars
                        FROM public.dwjjob j
                        LEFT JOIN public.dwvveh v ON j.dwjload = v.dwvload
                        WHERE j.dwjdate BETWEEN %s AND %s
                        GROUP BY j.dwjload, j.dwjcust
                    )
                    SELECT DISTINCT
                        l.load_date,
                        l.contractor,
                        l.collections,
                        l.deliveries,
                        l.total_cars,
                        h.start_time,
                        h.finish_time,
                        h.total_hours,
                        l.dwjload
                    FROM load_dates l
                    LEFT JOIN public.hours h ON CAST(l.load_date AS TEXT) = TO_CHAR(h.work_date, 'YYYYMMDD')
                    ORDER BY l.load_date, l.dwjload
                """, (start_date_str, end_date_str))
                
                loads = cursor.fetchall()
                
                if not loads:
                    print(f"{Fore.YELLOW}No loads found for this week.{Style.RESET_ALL}")
                    return False
                
                # Log detailed information about each load
                print(f"\n{Fore.CYAN}Detailed Load Information:{Style.RESET_ALL}")
                for i, load in enumerate(loads, 1):
                    date = datetime.strptime(str(load[0]), "%Y%m%d")
                    day_name = date.strftime("%A").upper()
                    print(f"\n{Fore.YELLOW}Load {i}:{Style.RESET_ALL}")
                    print(f"Load Number: {load[8]}")
                    print(f"Date: {date.strftime('%A %d-%m-%Y')}")
                    print(f"Contractor: {load[1]}")
                    print(f"Collections: {load[2]}")
                    print(f"Deliveries: {load[3]}")
                    print(f"Total Cars: {load[4]}")
                    print(f"Start Time: {load[5]}")
                    print(f"Finish Time: {load[6]}")
                    print(f"Total Hours: {load[7]}")
                    
                    logging.info(f"Load {i}: Date={date.strftime('%A %d-%m-%Y')}, Contractor={load[1]}, Cars={load[4]}, Hours={load[7]}")
                
                # Organize loads by day of week
                daily_loads = {}
                for load in loads:
                    date = datetime.strptime(str(load[0]), "%Y%m%d")
                    day_name = date.strftime("%A").upper()
                    if day_name not in daily_loads:
                        daily_loads[day_name] = {
                            'loads': [],
                            'start_time': load[5],
                            'finish_time': load[6],
                            'total_hours': load[7]
                        }
                    daily_loads[day_name]['loads'].append(load)
                
                logging.info(f"Organized loads by day: {list(daily_loads.keys())}")
                
                # Process each day's loads
                day_row_mapping = {
                    'MONDAY': {'start': 8, 'end': 10},
                    'TUESDAY': {'start': 11, 'end': 13},
                    'WEDNESDAY': {'start': 14, 'end': 16},
                    'THURSDAY': {'start': 17, 'end': 19},
                    'FRIDAY': {'start': 20, 'end': 22},
                    'SATURDAY': {'start': 23, 'end': 25},
                    'SUNDAY': {'start': 26, 'end': 28}
                }
                
                overflow_row = 29
                has_overflow = False
                
                # Define day order
                day_order = ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY', 'SUNDAY']
                
                for day in day_order:
                    if day in daily_loads:
                        day_data = daily_loads[day]
                        rows = day_row_mapping[day]
                        current_row = rows['start']
                        
                        logging.info(f"Processing {day}: {len(day_data['loads'])} loads")
                        
                        # Write loads for the day
                        for load in day_data['loads']:
                            if current_row <= rows['end']:
                                # Write contractor name
                                safe_cell_write(f'C{current_row}', load[1])
                                
                                # Write number of cars
                                safe_cell_write(f'D{current_row}', load[4])
                                
                                # Write collection towns
                                safe_cell_write(f'E{current_row}', load[2] or '')
                                
                                # Write delivery towns
                                safe_cell_write(f'F{current_row}', load[3] or '')
                                
                                logging.info(f"Wrote load data to row {current_row}: Contractor={load[1]}, Cars={load[4]}")
                                
                                current_row += 1
                            else:
                                # Mark that we have overflow loads
                                has_overflow = True
                                # Write overflow loads
                                safe_cell_write(f'C{overflow_row}', load[1])
                                safe_cell_write(f'D{overflow_row}', load[4])
                                safe_cell_write(f'E{overflow_row}', load[2] or '')
                                safe_cell_write(f'F{overflow_row}', load[3] or '')
                                logging.info(f"Wrote overflow load to row {overflow_row}")
                                overflow_row += 1
                        
                        # Write hours for the day if available
                        if day_data['start_time'] and day_data['finish_time']:
                            # Write hours in the first row of each day
                            safe_cell_write(f'H{rows["start"]}', format_time(day_data['start_time']))
                            safe_cell_write(f'I{rows["start"]}', format_time(day_data['finish_time']))
                            safe_cell_write(f'J{rows["start"]}', format_total_hours(day_data['total_hours']))
                            logging.info(f"Wrote hours for {day}: Start={day_data['start_time']}, Finish={day_data['finish_time']}, Total={day_data['total_hours']}")
                
                # Calculate and write total hours
                total_hours = sum(float(day_data['total_hours'] or 0) for day_data in daily_loads.values())
                safe_cell_write('J29', format_total_hours(total_hours))
                logging.info(f"Wrote total hours: {total_hours}")
                
                # Save the workbook
                workbook.save(output_file)
                logging.info(f"Successfully saved timesheet to {output_file}")
                
                # Verify the file was saved correctly
                if os.path.exists(output_file) and os.path.getsize(output_file) > 0:
                    print(f"{Fore.GREEN}Timesheet created successfully at {output_file}{Style.RESET_ALL}")
                    return True
                else:
                    print(f"{Fore.RED}Error: Timesheet file was not created or is empty{Style.RESET_ALL}")
                    return False
            finally:
                if 'cursor' in locals():
                    cursor.close()
                if 'conn' in locals():
                    conn.close()
            
        except Exception as e:
            logging.error(f"Error creating timesheet: {str(e)}")
            print(f"{Fore.RED}Error creating timesheet: {str(e)}{Style.RESET_ALL}")
            return False
        finally:
            if 'cursor' in locals():
                cursor.close()
            if 'conn' in locals():
                conn.close()

    def create_all_paperwork(self, selected_sunday):
        """Create all loadsheets and timesheet for the selected week."""
        try:
            # Calculate week start (Monday)
            week_start = selected_sunday - timedelta(days=6)
            week_end = selected_sunday
            
            print(f"\n{Fore.CYAN}Creating paperwork for week:{Style.RESET_ALL}")
            print(f"{Fore.WHITE}Week Start (Monday): {Fore.YELLOW}{week_start.strftime('%A %d-%m-%Y')}{Style.RESET_ALL}")
            print(f"{Fore.WHITE}Week End (Sunday): {Fore.YELLOW}{week_end.strftime('%A %d-%m-%Y')}{Style.RESET_ALL}")
            
            # Get all loads for the week
            loads = self.get_loads_for_week(selected_sunday)
            if not loads:
                print(f"{Fore.YELLOW}No loads found for this week.{Style.RESET_ALL}")
                return False
                
            print(f"\n{Fore.CYAN}Found {len(loads)} loads for this week:{Style.RESET_ALL}")
            for load in loads:
                print(f"{Fore.WHITE}Load: {Fore.YELLOW}{load[0]}{Style.RESET_ALL}")
                
            confirm = input(f"\n{Fore.YELLOW}Create all paperwork for this week? (y/n):{Style.RESET_ALL} ").strip().lower()
            if confirm != 'y':
                print(f"{Fore.YELLOW}Operation cancelled.{Style.RESET_ALL}")
                return False
            
            # Create base directories
            loadsheets_dir = os.path.join(SCRIPT_DIR, "loadsheets")
            timesheets_dir = os.path.join(SCRIPT_DIR, "timesheets")
            os.makedirs(loadsheets_dir, exist_ok=True)
            os.makedirs(timesheets_dir, exist_ok=True)
            
            # Create week-specific directories
            week_folder = os.path.join(loadsheets_dir, week_end.strftime("%d-%m-%Y"))
            timesheet_folder = os.path.join(timesheets_dir, week_end.strftime("%Y%m%d"))
            os.makedirs(week_folder, exist_ok=True)
            os.makedirs(timesheet_folder, exist_ok=True)
            
            print(f"\n{Fore.CYAN}Creating paperwork in:{Style.RESET_ALL}")
            print(f"{Fore.WHITE}Loadsheets: {Fore.YELLOW}{week_folder}{Style.RESET_ALL}")
            print(f"{Fore.WHITE}Timesheet: {Fore.YELLOW}{timesheet_folder}{Style.RESET_ALL}")
            
            # Create timesheet first
            print(f"\n{Fore.CYAN}Creating timesheet...{Style.RESET_ALL}")
            if not self.create_timesheet(selected_sunday):
                print(f"{Fore.RED}Failed to create timesheet. Stopping paperwork creation.{Style.RESET_ALL}")
                return False
            
            # Create loadsheets for each load
            print(f"\n{Fore.CYAN}Creating loadsheets...{Style.RESET_ALL}")
            for i, load in enumerate(loads, 1):
                print(f"\n{Fore.YELLOW}Processing load {i} of {len(loads)}: {load[0]}{Style.RESET_ALL}")
                
                try:
                    # Use the existing create_loadsheet method
                    if not self.create_loadsheet(load[0]):
                        print(f"{Fore.RED}Failed to create loadsheet for load {load[0]}. Skipping...{Style.RESET_ALL}")
                        continue
                    
                except Exception as e:
                    print(f"{Fore.RED}Error creating loadsheet for load {load[0]}: {e}{Style.RESET_ALL}")
                    logging.error(f"Error creating loadsheet for load {load[0]}: {e}", exc_info=True)
                    continue
            
            print(f"\n{Fore.GREEN}All paperwork created successfully!{Style.RESET_ALL}")
            return True
            
        except Exception as e:
            logging.error(f"Error in create_all_paperwork: {e}", exc_info=True)
            print(f"{Fore.RED}Error creating paperwork: {e}{Style.RESET_ALL}")
            return False

    def check_required_files(self):
        """Check if all required files and directories exist."""
        # Check template files
        template_files = {
            'templates/loadsheet.xlsx': os.path.join(SCRIPT_DIR, "templates", "loadsheet.xlsx"),
            'templates/timesheet.xlsx': os.path.join(SCRIPT_DIR, "templates", "timesheet.xlsx")
        }
        
        # Check signature directories
        signature_dirs = {
            'signature/sig1': os.path.join(SCRIPT_DIR, "signature", "sig1"),
            'signature/sig2': os.path.join(SCRIPT_DIR, "signature", "sig2")
        }
        
        missing_files = []
        missing_dirs = []
        
        # Check template files
        for file_desc, file_path in template_files.items():
            if not os.path.exists(file_path):
                missing_files.append(file_desc)
        
        # Check signature directories and their contents
        for dir_desc, dir_path in signature_dirs.items():
            if not os.path.exists(dir_path):
                missing_dirs.append(dir_desc)
            else:
                # Count PNG files in the directory
                png_files = [f for f in os.listdir(dir_path) if f.lower().endswith('.png')]
                if len(png_files) < 1:
                    missing_files.append(f"{dir_desc}/*.png (no signature files found)")
        
        # Report missing items
        if missing_files or missing_dirs:
            print(f"\n{Fore.RED}Missing required files and directories:{Style.RESET_ALL}")
            if missing_dirs:
                print(f"\n{Fore.YELLOW}Missing directories:{Style.RESET_ALL}")
                for dir_name in missing_dirs:
                    print(f"{Fore.YELLOW}- {dir_name}{Style.RESET_ALL}")
            if missing_files:
                print(f"\n{Fore.YELLOW}Missing files:{Style.RESET_ALL}")
                for file_name in missing_files:
                    print(f"{Fore.YELLOW}- {file_name}{Style.RESET_ALL}")
            return False
            
        print(f"{Fore.GREEN}All required files and directories are present.{Style.RESET_ALL}")
        return True

    def generate_load_summary(self, vehicles):
        """Generate a summary message for the loadsheet."""
        total_cars = len(vehicles)
        offloaded_cars = sum(1 for v in vehicles if v[2] == 'Y')  # Assuming v[2] is offloaded status
        loaded_cars = total_cars - offloaded_cars
        
        # Count cars with documents and spare keys
        cars_with_docs = sum(1 for v in vehicles if v[3] == 'Y')  # Assuming v[3] is documents status
        cars_with_keys = sum(1 for v in vehicles if v[4] == 'Y')  # Assuming v[4] is spare keys status
        
        # Build the message
        message_parts = []
        
        if offloaded_cars > 0:
            message_parts.append(f"{offloaded_cars} CAR{'S' if offloaded_cars > 1 else ''} OFFLOADED")
        
        if loaded_cars > 0:
            message_parts.append(f"{loaded_cars} CAR{'S' if loaded_cars > 1 else ''} LOADED")
        
        if cars_with_docs > 0:
            message_parts.append(f"{cars_with_docs} CAR{'S' if cars_with_docs > 1 else ''} HAVE DOCUMENTS ALL DOCUMENTS ON PASSENGER SEAT'S")
        
        if cars_with_keys > 0:
            message_parts.append(f"{cars_with_keys} CAR{'S' if cars_with_keys > 1 else ''} HAVE SPARE KEYS")
        
        return ", ".join(message_parts)

    def add_signatures(self, ws):
        """Add signatures to the loadsheet with fine-tuned positioning."""
        try:
            # Only add signatures to loadsheets
            if ws.title != "Loadsheet":
                return
                
            # Get signature files
            sig1_dir = os.path.join(SCRIPT_DIR, "signature", "sig1")
            sig2_dir = os.path.join(SCRIPT_DIR, "signature", "sig2")
            
            sig1_files = [f for f in os.listdir(sig1_dir) if f.lower().endswith('.png')]
            sig2_files = [f for f in os.listdir(sig2_dir) if f.lower().endswith('.png')]
            
            if not sig1_files or not sig2_files:
                print(f"{Fore.YELLOW}Warning: Missing signature files{Style.RESET_ALL}")
                logging.error("No signature files found in one or both directories")
                return
            
            # Randomly select signature files
            sig1_path = os.path.join(sig1_dir, random.choice(sig1_files))
            sig2_files = [f for f in sig2_files if f != os.path.basename(sig1_path)]
            if not sig2_files:
                sig2_files = [f for f in os.listdir(sig2_dir) if f.lower().endswith('.png')]
            sig2_path = os.path.join(sig2_dir, random.choice(sig2_files))
            
            # Create images
            img1 = OpenpyxlImage(sig1_path)
            img2 = OpenpyxlImage(sig2_path)
            
            # Set base size with 1.5x scaling
            base_width = 100
            base_height = 50
            scale_factor = 1.5
            
            img1.width = int(base_width * scale_factor)
            img1.height = int(base_height * scale_factor)
            img2.width = int(base_width * scale_factor)
            img2.height = int(base_height * scale_factor)
            
            # Get cell positions for loadsheet
            sig1_cell = 'C44'
            sig2_cell = 'H44'
            
            # Get cell positions and dimensions
            x1, y1 = self.get_cell_position(ws, sig1_cell)
            x2, y2 = self.get_cell_position(ws, sig2_cell)
            
            width1, height1 = self.get_cell_dimensions(ws, sig1_cell)
            width2, height2 = self.get_cell_dimensions(ws, sig2_cell)
            
            # Randomly choose a vertical offset mode (4-7)
            vertical_mode = random.randint(4, 7)
            
            # Add fixed right offset for sig1
            SIG1_RIGHT_OFFSET = 35  # Fixed right offset for sig1
            x1 = x1 + SIG1_RIGHT_OFFSET
            
            # Generate random rotation (-15 to +15 degrees)
            rotation = random.randint(-15, 15)
            
            # Apply placement mode
            if vertical_mode == 4:  # Centered with small vertical offset
                x1_centered = x1 + (width1 - img1.width) / 2
                y1_centered = y1 + (height1 - img1.height) / 2 - 10
                x2_centered = x2 + (width2 - img2.width) / 2
                y2_centered = y2 + (height2 - img2.height) / 2 - 10
                
                img1.anchor = self.create_absolute_anchor(x1_centered, y1_centered, img1.width, img1.height)
                img2.anchor = self.create_absolute_anchor(x2_centered, y2_centered, img2.width, img2.height)
            
            elif vertical_mode == 5:  # Basic with medium vertical offset
                y1_offset = y1 - 20
                y2_offset = y2 - 20
                
                img1.anchor = self.create_absolute_anchor(x1, y1_offset, img1.width, img1.height)
                img2.anchor = self.create_absolute_anchor(x2, y2_offset, img2.width, img2.height)
            
            elif vertical_mode == 6:  # Centered with medium vertical offset
                x1_centered = x1 + (width1 - img1.width) / 2
                y1_centered = y1 + (height1 - img1.height) / 2 - 20
                x2_centered = x2 + (width2 - img2.width) / 2
                y2_centered = y2 + (height2 - img2.height) / 2 - 20
                
                img1.anchor = self.create_absolute_anchor(x1_centered, y1_centered, img1.width, img1.height)
                img2.anchor = self.create_absolute_anchor(x2_centered, y2_centered, img2.width, img2.height)
            
            elif vertical_mode == 7:  # Basic with large vertical offset
                y1_offset = y1 - 30
                y2_offset = y2 - 30
                
                img1.anchor = self.create_absolute_anchor(x1, y1_offset, img1.width, img1.height)
                img2.anchor = self.create_absolute_anchor(x2, y2_offset, img2.width, img2.height)
            
            # Add rotation to both images
            img1.rotation = rotation
            img2.rotation = rotation
            
            # Add images to worksheet
            ws.add_image(img1)
            ws.add_image(img2)
            
            print(f"{Fore.GREEN}Signatures added successfully with rotation {rotation}°{Style.RESET_ALL}")
            
        except Exception as e:
            print(f"{Fore.YELLOW}Warning: Could not add signatures: {e}{Style.RESET_ALL}")
            logging.error(f"Error adding signatures: {e}", exc_info=True)

    def get_cell_position(self, ws, cell):
        """Get cell position in pixels from top-left of sheet."""
        col, row = coordinate_from_string(cell)
        col_idx = column_index_from_string(col)
        
        # Calculate position by summing up previous column widths and row heights
        x = 0
        for i in range(1, col_idx):
            col_letter = chr(ord('A') + i - 1)
            col_width = ws.column_dimensions[col_letter].width or 8.43
            x += col_width * 7
        
        y = 0
        for i in range(1, row):
            row_height = ws.row_dimensions[i].height or 15
            y += row_height * 1.2
            
        # Add vertical offset to move down from row 44
        VERTICAL_ROW_OFFSET = 18 * 7  # Move down 7 rows from row 44
        y = y + VERTICAL_ROW_OFFSET
            
        return x, y

    def get_cell_dimensions(self, ws, cell):
        """Get cell dimensions in pixels."""
        col, row = coordinate_from_string(cell)
        col_idx = column_index_from_string(col)
        
        # Get column width (in characters)
        col_width = ws.column_dimensions[col].width or 8.43
        # Get row height (in points)
        row_height = ws.row_dimensions[row].height or 15
        
        # Convert to pixels
        width_px = col_width * 7
        height_px = row_height * 1.2
        
        return width_px, height_px

    def create_absolute_anchor(self, x_px, y_px, width_px, height_px):
        """Create an absolute anchor with size."""
        EMU_PER_PIXEL = 9525
        x_emu = int(x_px * EMU_PER_PIXEL)
        y_emu = int(y_px * EMU_PER_PIXEL)
        width_emu = int(width_px * EMU_PER_PIXEL)
        height_emu = int(height_px * EMU_PER_PIXEL)
        
        anchor = AbsoluteAnchor(pos=XDRPoint2D(x=x_emu, y=y_emu))
        anchor.ext = XDRPositiveSize2D(cx=width_emu, cy=height_emu)
        return anchor

    def run(self):
        """Main menu loop."""
        while True:
            print_header()
            print_menu()
            
            choice = input(f"\n{Fore.CYAN}Enter your choice (1-5):{Style.RESET_ALL} ").strip()
            
            if choice == "1":
                print_status("Creating all paperwork...")
                selected_sunday = self.select_week()
                if selected_sunday:
                    if self.create_all_paperwork(selected_sunday):
                        print_status("All paperwork created successfully!", "success")
                    else:
                        print_status("Failed to create paperwork.", "error")
            
            elif choice == "2":
                print_status("Creating single loadsheet...")
                load_number = input(f"{Fore.CYAN}Enter load number:{Style.RESET_ALL} ").strip()
                if load_number:
                    if self.create_loadsheet(load_number):
                        print_status("Loadsheet created successfully!", "success")
                    else:
                        print_status("Failed to create loadsheet.", "error")
            
            elif choice == "3":
                print_status("Creating timesheet...")
                selected_sunday = self.select_week()
                if selected_sunday:
                    if self.create_timesheet(selected_sunday):
                        print_status("Timesheet created successfully!", "success")
                    else:
                        print_status("Failed to create timesheet.", "error")
            
            elif choice == "4":
                print_status("Toggling auto signature...")
                if self.toggle_auto_signature():
                    print_status("Auto signature toggled successfully!", "success")
                else:
                    print_status("Failed to toggle auto signature.", "error")
            
            elif choice == "5":
                print_status("Exiting program...", "info")
                break
            
            else:
                print_status("Invalid choice. Please try again.", "error")
            
            input(f"\n{Fore.YELLOW}Press Enter to continue...{Style.RESET_ALL}")

if __name__ == "__main__":
    manager = PaperworkManager()
    try:
        manager.run()
    except KeyboardInterrupt:
        print(f"\n{Fore.YELLOW}Operation interrupted. Exiting...{Style.RESET_ALL}")
    except Exception as e:
        print(f"{Fore.RED}An unexpected error occurred: {e}{Style.RESET_ALL}")
    finally:
        print(f"{Fore.GREEN}Program terminated.{Style.RESET_ALL}") 